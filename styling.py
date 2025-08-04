from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Farbkonfiguration
HEADER_FILL_COLOR = "E20074"
HEADER_FONT_COLOR = "FFFFFF"
ABWEICHUNG_POSITIV_ROT = "FFC7CE"
ABWEICHUNG_NEGATIV_GRUEN = "C6EFCE"
ZEBRA_ROW_COLOR = "F2F2F2"
BESTBIETER_GRUEN = "C6EFCE"


def style_auswertungs_sheet(ws, df_columns):
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    header_font = Font(color=HEADER_FONT_COLOR, bold=True, size=12, name="Calibri")
    cell_font = Font(size=10, name="Calibri")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    fill_rot = PatternFill(start_color=ABWEICHUNG_POSITIV_ROT, end_color=ABWEICHUNG_POSITIV_ROT, fill_type="solid")
    fill_gruen = PatternFill(start_color=ABWEICHUNG_NEGATIV_GRUEN, end_color=ABWEICHUNG_NEGATIV_GRUEN,
                             fill_type="solid")
    zebra_fill = PatternFill(start_color=ZEBRA_ROW_COLOR, end_color=ZEBRA_ROW_COLOR, fill_type="solid")

    for idx, title in enumerate(df_columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    ws.row_dimensions[1].height = 22

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max_len + 4

    header_map = {title: i + 1 for i, title in enumerate(df_columns)}

    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = zebra_fill

        for col_title, col_index in header_map.items():
            cell = ws.cell(row=row, column=col_index)
            cell.font = cell_font
            cell.alignment = right_align if col_title.endswith("€") or "Abweichung (%)" in col_title else center_align

            if col_title in ["Gesamtpreis (€)", "Abweichung (€)", "Banf (€)"]:
                cell.number_format = '€ #,##0.00'

            if col_title == "Abweichung (%)":
                cell.number_format = '0.00%'
                val = cell.value
                if isinstance(val, (int, float)):
                    if val > 0:
                        cell.fill = fill_rot
                    elif val < 0:
                        cell.fill = fill_gruen

    ws.auto_filter.ref = ws.dimensions


def style_pivot_sheet(ws, df_columns):
    """
    Formatiert das Pivot-Sheet 'Vergleich':
    - Euro-Format für Preise
    - Rechtsbündig für Zahlen
    - Zentrierter Header
    - Filter aktiv
    - Auto-Spaltenbreiten
    """
    euro_format = '€ #,##0.00'
    header_font = Font(bold=True, size=12, name="Calibri", color="FFFFFF")
    body_font = Font(name="Calibri", size=10)
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="E20074", end_color="E20074", fill_type="solid")

    # Header formatieren
    for idx, title in enumerate(df_columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Zellen formatieren
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):  # ab Spalte 2 = Zahlen (Preise)
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if isinstance(val, (int, float)):
                cell.number_format = euro_format
                cell.font = body_font
                cell.alignment = right_align

    # Spaltenbreiten automatisch
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max_len + 4

    # Filter aktivieren (z. B. auf "Los")
    ws.auto_filter.ref = ws.dimensions